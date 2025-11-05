"""
Excel document processing module - データ、表、画像の抽出
"""

import logging
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import tiktoken
from PIL import Image
import io
from langchain_text_splitters import RecursiveCharacterTextSplitter


logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Excelファイルからデータ、表、画像を抽出するクラス"""

    def __init__(self, config: Dict[str, Any]):
        """
        初期化

        Args:
            config: 処理設定
        """
        self.config = config
        self.pdf_config = config.get("pdf_processing", {})  # 同じ設定を使用
        self.vision_config = config.get("vision", {})
        self.rag_config = config.get("rag", {})
        self.chunking_config = config.get("chunking", {})

        # tiktokenエンコーダーの初期化
        model_name = config.get("openai", {}).get("model_chat", "gpt-4.1")
        try:
            self.encoding = tiktoken.encoding_for_model(model_name)
        except KeyError:
            self.encoding = tiktoken.get_encoding("cl100k_base")

        # セマンティックチャンカーの初期化
        self.text_splitter = None
        if self.rag_config.get("enable_semantic_chunking", False):
            try:
                chunk_size = self.chunking_config.get("chunk_size", self.pdf_config.get("chunk_size", 800))
                chunk_overlap = self.chunking_config.get("chunk_overlap", self.pdf_config.get("chunk_overlap", 150))
                separators = self.chunking_config.get("separators", ["\n\n", "\n", "。", "．", ". ", "! ", "? ", "；", "、", ", ", " ", ""])

                self.text_splitter = RecursiveCharacterTextSplitter(
                    chunk_size=chunk_size,
                    chunk_overlap=chunk_overlap,
                    length_function=lambda text: len(self.encoding.encode(text)),
                    separators=separators,
                    keep_separator=True
                )
                logger.info(f"Semantic chunking initialized for Excel (chunk_size={chunk_size}, overlap={chunk_overlap})")
            except Exception as e:
                logger.error(f"Failed to initialize semantic chunker: {e}")
                logger.warning("Falling back to legacy chunking")

    def process_excel(self, excel_path: str, category: str) -> Dict[str, Any]:
        """
        Excelファイルを処理してデータ、表、画像を抽出

        Args:
            excel_path: Excelファイルのパス
            category: ドキュメントカテゴリー

        Returns:
            dict: 抽出結果（テキストチャンク、画像情報など）
        """
        logger.info(f"Processing Excel document: {excel_path}")

        result = {
            "source_file": Path(excel_path).name,
            "category": category,
            "text_chunks": [],
            "images": [],
            "metadata": {},
        }

        try:
            # Excelファイルを開く
            wb = openpyxl.load_workbook(excel_path, data_only=True)  # data_only=True: 数式の結果を取得

            # メタデータ
            result["metadata"]["sheets_count"] = len(wb.sheetnames)
            result["metadata"]["sheets"] = wb.sheetnames

            # 画像出力ディレクトリを準備
            doc_name = Path(excel_path).stem
            output_dir = Path(f"data/extracted_images/{doc_name}")
            output_dir.mkdir(parents=True, exist_ok=True)

            # 各シートを処理
            for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
                logger.info(f"Processing sheet {sheet_index}/{len(wb.sheetnames)}: {sheet_name}")

                sheet = wb[sheet_name]
                sheet_text = []

                # シート名をヘッダーとして追加
                sheet_text.append(f"\n\n## シート: {sheet_name}\n\n")

                # データ範囲を取得
                if sheet.max_row > 0 and sheet.max_column > 0:
                    # 表データをMarkdownに変換
                    table_markdown = self._sheet_to_markdown(sheet)

                    if table_markdown:
                        sheet_text.append(table_markdown)
                        sheet_text.append("\n\n")

                # シートのテキストを結合
                full_sheet_text = "".join(sheet_text)

                # テキストをチャンクに分割
                if full_sheet_text.strip():
                    chunks = self._create_text_chunks(
                        full_sheet_text,
                        section_num=sheet_index,  # シート番号をセクション番号として使用
                        source_file=result["source_file"],
                        category=category,
                        sheet_name=sheet_name
                    )
                    result["text_chunks"].extend(chunks)

                # 画像を抽出（openpyxlは画像抽出をサポートしているが、_imagesは内部属性）
                if hasattr(sheet, '_images'):
                    for img_index, img in enumerate(sheet._images):
                        try:
                            # 画像データを取得
                            image_data = img.ref.getvalue()

                            # PIL Imageに変換
                            pil_image = Image.open(io.BytesIO(image_data))

                            # 画像をリサイズ（大きすぎる場合）
                            max_size = self.vision_config.get("max_image_size", 2000)
                            if max(pil_image.size) > max_size:
                                ratio = max_size / max(pil_image.size)
                                new_size = tuple(int(dim * ratio) for dim in pil_image.size)
                                pil_image = pil_image.resize(new_size, Image.Resampling.LANCZOS)

                            # 画像を保存
                            image_filename = f"sheet_{sheet_index}_image_{img_index}.png"
                            image_path = output_dir / image_filename
                            pil_image.save(str(image_path), format=self.vision_config.get("image_format", "PNG"))

                            # メタデータを作成
                            image_info = {
                                "image_path": str(image_path),
                                "page_number": sheet_index,  # シート番号
                                "image_index": img_index,
                                "width": pil_image.size[0],
                                "height": pil_image.size[1],
                                "source_file": result["source_file"],
                                "category": category,
                                "content_type": "image",
                                "sheet_name": sheet_name
                            }
                            result["images"].append(image_info)

                            logger.debug(f"Extracted image {img_index} from sheet {sheet_name}")

                        except Exception as e:
                            logger.warning(f"Failed to extract image {img_index} from sheet {sheet_name}: {e}")
                            continue

            logger.info(f"Extracted {len(result['text_chunks'])} text chunks and {len(result['images'])} images from Excel document")

        except Exception as e:
            logger.error(f"Error processing Excel document: {e}")
            raise

        return result

    def _sheet_to_markdown(self, sheet: Worksheet, max_rows: int = 1000, max_cols: int = 50) -> str:
        """
        ExcelシートをMarkdown形式の表に変換

        Args:
            sheet: openpyxlのWorksheetオブジェクト
            max_rows: 処理する最大行数（大きすぎるシートを避けるため）
            max_cols: 処理する最大列数

        Returns:
            str: Markdown形式の表
        """
        if sheet.max_row == 0 or sheet.max_column == 0:
            return ""

        # 実際の行数と列数を制限
        actual_rows = min(sheet.max_row, max_rows)
        actual_cols = min(sheet.max_column, max_cols)

        if actual_rows < sheet.max_row or actual_cols < sheet.max_column:
            logger.warning(f"Sheet too large, limiting to {actual_rows} rows and {actual_cols} columns")

        markdown_lines = []

        # データを読み込む
        data = []
        for row_idx in range(1, actual_rows + 1):
            row_data = []
            for col_idx in range(1, actual_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value

                # Noneや空白を扱う
                if value is None:
                    row_data.append("")
                else:
                    # 数式の場合は結果を、それ以外は値をそのまま
                    row_data.append(str(value).strip())

            # 空行をスキップ（全てのセルが空の場合）
            if any(cell for cell in row_data):
                data.append(row_data)

        if not data:
            return ""

        # Markdownテーブルを構築
        # 1行目をヘッダーとして扱う
        if data:
            header = "| " + " | ".join(data[0]) + " |"
            markdown_lines.append(header)

            # セパレーター
            separator = "| " + " | ".join(["---"] * len(data[0])) + " |"
            markdown_lines.append(separator)

            # データ行
            for row in data[1:]:
                # 列数を揃える（ヘッダーと同じ列数になるようにパディング）
                while len(row) < len(data[0]):
                    row.append("")

                data_row = "| " + " | ".join(row[:len(data[0])]) + " |"
                markdown_lines.append(data_row)

        return "\n".join(markdown_lines)

    def _create_text_chunks(
        self,
        text: str,
        section_num: int,
        source_file: str,
        category: str,
        sheet_name: str = None
    ) -> List[Dict[str, Any]]:
        """
        テキストをチャンクに分割

        Args:
            text: テキスト
            section_num: セクション番号（シート番号）
            source_file: ソースファイル名
            category: カテゴリー
            sheet_name: シート名

        Returns:
            list: テキストチャンクのリスト
        """
        # セマンティックチャンキング or レガシーチャンキング
        if self.text_splitter:
            # セマンティックチャンキング使用
            chunk_texts = self.text_splitter.split_text(text)
            chunks = []
            for idx, chunk_text in enumerate(chunk_texts):
                if chunk_text.strip():
                    token_count = len(self.encoding.encode(chunk_text))
                    chunk_data = {
                        "text": chunk_text.strip(),
                        "page_number": section_num,  # シート番号
                        "source_file": source_file,
                        "category": category,
                        "chunk_index": idx,
                        "token_count": token_count,
                    }
                    if sheet_name:
                        chunk_data["sheet_name"] = sheet_name
                    chunks.append(chunk_data)
            logger.debug(f"Created {len(chunks)} semantic chunks from Excel sheet")
            return chunks
        else:
            # レガシーチャンキング
            return self._legacy_create_text_chunks(text, section_num, source_file, category, sheet_name)

    def _legacy_create_text_chunks(
        self,
        text: str,
        section_num: int,
        source_file: str,
        category: str,
        sheet_name: str = None
    ) -> List[Dict[str, Any]]:
        """
        テキストをチャンクに分割（レガシー実装）

        Args:
            text: テキスト
            section_num: セクション番号
            source_file: ソースファイル名
            category: カテゴリー
            sheet_name: シート名

        Returns:
            list: テキストチャンクのリスト
        """
        chunk_size = self.pdf_config.get("chunk_size", 800)
        chunk_overlap = self.pdf_config.get("chunk_overlap", 150)
        chunks = []

        # テキストをトークン化
        tokens = self.encoding.encode(text)

        # トークン数が少ない場合はそのまま返す
        if len(tokens) <= chunk_size:
            chunk_data = {
                "text": text.strip(),
                "page_number": section_num,
                "source_file": source_file,
                "category": category,
                "chunk_index": 0,
                "token_count": len(tokens)
            }
            if sheet_name:
                chunk_data["sheet_name"] = sheet_name
            return [chunk_data]

        # チャンクに分割
        start = 0
        chunk_index = 0

        while start < len(tokens):
            end = min(start + chunk_size, len(tokens))
            chunk_tokens = tokens[start:end]
            chunk_text = self.encoding.decode(chunk_tokens)

            if chunk_text.strip():
                chunk_data = {
                    "text": chunk_text.strip(),
                    "page_number": section_num,
                    "source_file": source_file,
                    "category": category,
                    "chunk_index": chunk_index,
                    "token_count": len(chunk_tokens),
                }
                if sheet_name:
                    chunk_data["sheet_name"] = sheet_name
                chunks.append(chunk_data)
                chunk_index += 1

            start = end - chunk_overlap if end < len(tokens) else end

        logger.debug(f"Created {len(chunks)} chunks from Excel sheet ({len(tokens)} tokens)")
        return chunks
